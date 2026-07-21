# Prepare dataset in .xlsx file for enrichment:
# - remove chains with small amount of samples (based on initial analysis)
# - extract contract address for further enrichment with on-chain features
# - check dataset for identical rows by contract address, print rows that require manual check before proceeding

# OpenPyXL is used here to preserve hyperlinks in the file.
# All actions with .xlsx are based on initial analysis (see Jupyter notebook "TM-RugPull initial analysis").
# They are grouped here for convenience (to edit .xlxs file once and save results to a new ready-to-go file)

# Reference: https://openpyxl.readthedocs.io/en/3.1/tutorial.html

from openpyxl import load_workbook, Workbook
import re


# Load .xlsx file using OpenPyXL
# Returns workbook and sheet from .xlcx file
def load_file (file):
    workbook = load_workbook(file)
    sheet = workbook['Sheet1']
    return workbook, sheet


# Drop rows where 'Blockchain' column contains value for networks that were excluded from further analysis
# Returns a new re-created workbook to avoid a bug with preserving old hyperlinks when shifting rows due to deletion
def drop_chains(sheet, headings, chains_to_drop):
    blockchain_column_idx = headings.index('Blockchain')

    # Create new workbook with new sheet
    workbook_to_proceed = Workbook()
    sheet_to_proceed = workbook_to_proceed.active
    sheet_to_proceed.title = sheet.title

    # Copy header row to new workbook
    for col_idx, value in enumerate(headings, start=1):
        sheet_to_proceed.cell(row=1, column=col_idx, value=value)

    # Copy all rows that should remain
    new_row_number = 2
    for row in sheet.iter_rows(min_row=2):
        if row[blockchain_column_idx].value in chains_to_drop:
            continue
        for col_idx, cell in enumerate(row, start=1):
            new_cell = sheet_to_proceed.cell(row=new_row_number, column=col_idx, value=cell.value)
            if cell.hyperlink is not None:
                new_cell.hyperlink = cell.hyperlink.target
        new_row_number += 1

    # Check how many rows remain (should be 990 for TM-RugPull dataset)
    print(new_row_number - 2, "rows remaining")
    return workbook_to_proceed, sheet_to_proceed


# Add contract addresses to all remaining rows in the file for further extraction of on-chain features
# Addresses are extracted from hyperlinks from 'Smart Contract (online)' column
# Reference: https://docs.python.org/3/library/re.html
def add_contract_addresses(sheet, headings):
    contract_link_column_idx = headings.index('Smart Contract (online)')
    address_column_number = len(headings) + 1

    # A regex pattern for contract address that is used to extract addresses from URLs
    # Compiled once for efficiency
    address_pattern = re.compile(r'0x[a-fA-F0-9]{40}')

    sheet.cell(row=1, column=address_column_number, value='Contract address')

    for row in sheet.iter_rows(min_row=2):
        cell = row[contract_link_column_idx]
        if cell.hyperlink is None:
            print(f"Row {cell.row}: no hyperlink")
            continue
        # Reference: https://openpyxl.readthedocs.io/en/3.1/api/openpyxl.worksheet.hyperlink.html
        match = address_pattern.search(cell.hyperlink.target)
        if match is None:
            print(f"Row {cell.row}: could not extract address")
            continue
        address = match.group(0).lower()
        sheet.cell(row=cell.row, column=address_column_number, value=address)


# Check dataset for duplicates (rows with the same contract address)
def check_duplicates_by_address (sheet):
    headings = [c.value for c in sheet[1]]          # Repeated here, since a new heading ('Contract address') should be added on previous step
    address_column_idx = headings.index('Contract address')
    rows_per_address = {}
    for row in sheet.iter_rows(min_row=2):
        address = row[address_column_idx].value
        if address is None:
            continue
        rows_per_address.setdefault(address, []).append(row[0].row)

    for address, rows in rows_per_address.items():
        if len(rows) > 1:
            print(f"{rows} contain duplicated contract address {address}")


# Save a workbook with amenmdnments to a new .xlxs file
def save_workbook(workbook, output_file):
    workbook.save(output_file)


# Combine all actions with a file together
def main():
    workbook, sheet = load_file("data/TM-RugPull_original.xlsx")
    headings = [c.value for c in sheet[1]]
    # Based on initial analysis
    chains_to_drop = {'FANTOM', 'CRONO', 'BASE', 'FTM', 'SNOW'}
    new_workbook, new_sheet = drop_chains(sheet, headings, chains_to_drop)
    # Add contract addresses and check for duplicates based on addresses
    add_contract_addresses(new_sheet, headings)
    check_duplicates_by_address(new_sheet)
    save_workbook(new_workbook, 'data/TM-RugPull_prepared_for_enrichment.xlsx')


if __name__ == "__main__":
    main()
